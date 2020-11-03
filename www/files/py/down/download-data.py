#!/usr/bin/env python3
import re
import os
import string
import random
import requests
from json import dump
from datetime import datetime
from subprocess import call
from filecmp import cmp

from openpyxl import load_workbook

URL = CENSORED
EXT = 'xlsm'

NOW = datetime.now()
HOME_DIR = '/home/mushinako/www/files/'
XLSX_DIR = HOME_DIR + 'xlsx/'
RAND_LEN = 32
LOCK_DIR = HOME_DIR + 'key/'
LOCK = LOCK_DIR + 'down.lock'
UPD_TIME = LOCK_DIR + 'time.lock'
FP = LOCK_DIR + 'fp.lock'
DOWN_LOG = HOME_DIR + f'log/down/{NOW.strftime(r"%Y-%m-%d (%a)")}.log'
WORKSHEET = 'Events'
COL_START = 30
ROW_NAME = 11
SUMM = LOCK_DIR + 'summ.lock'


def random_string():
    cand = string.digits + string.ascii_letters
    return ''.join([random.choice(cand) for _ in range(RAND_LEN)])


def down():
    with open(LOCK, 'w') as f:
        f.write('1')
    file_url = requests.get(URL).content
    file_get_url = re.search(
        br'\"FileGetUrl\":\"(https://[A-Za-z0-9/_\-=\\\.\?]+)\",',
        file_url
    )[1]
    dl_url = file_get_url.replace(b'\\u0026', b'&')
    file = requests.get(dl_url).content
    new_len = len(file)
    with open(FP, 'r') as f:
        fp = f.read()
    act_xlsx = f'{XLSX_DIR}{random_string()}.{EXT}'
    tmp_xlsx = f'{XLSX_DIR}{random_string()}.{EXT}'
    with open(tmp_xlsx, 'wb') as f:
        f.write(file)
    os.replace(tmp_xlsx, act_xlsx)
    with open(FP, 'w') as f:
        f.write(act_xlsx)
    return (fp, act_xlsx, new_len)


def parse(fn):
    wb = load_workbook(fn, read_only=True)
    ws = wb[WORKSHEET]
    for row in ws.iter_rows(min_row=ROW_NAME, max_row=ROW_NAME, min_col=COL_START, values_only=True):
        col = 0
        names = {}
        while True:
            name = row[col]
            if name == '<Member Name>' or name is None:
                break
            real_col = COL_START + col
            names[name] = real_col
            col += 3
        break
    act_json = f'{XLSX_DIR}{random_string()}.json'
    tmp_json = f'{XLSX_DIR}{random_string()}.json'
    with open(tmp_json, 'w') as f:
        dump(names, f)
    os.replace(tmp_json, act_json)
    with open(SUMM, 'r') as f:
        summ = f.read()
    if os.path.isfile(summ):
        os.remove(summ)
    with open(SUMM, 'w') as f:
        f.write(act_json)
    with open(LOCK, 'w') as f:
        f.write('0')


def log(old_len, new_len):
    d = NOW.strftime('%Y-%m-%d (%a) %H:%M')
    with open(UPD_TIME, 'w') as f:
        f.write(d)
    with open(DOWN_LOG, 'a') as f:
        f.write(f'{d} {old_len} -> {new_len} ({new_len-old_len})\n')


def check(old_fn, new_fn):
    if not os.path.isfile(old_fn):
        return (False, 0)
    if cmp(old_fn, new_fn):
        os.remove(old_fn)
        return (True, 0)
    with open(old_fn, 'rb') as f:
        old_len = len(f.read())
    os.remove(old_fn)
    return (False, old_len)


def main():
    old_fn, new_fn, new_len = down()
    parse(new_fn)
    same, old_len = check(old_fn, new_fn)
    if not same:
        log(old_len, new_len)


if __name__ == '__main__':
    main()
